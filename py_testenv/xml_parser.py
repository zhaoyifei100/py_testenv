
"""
XMLParser
--------------------------------------
多功能XML解析与寄存器配置工具类。

功能：
1. 解析寄存器XML，提取设备地址映射。
2. 生成寄存器定义py文件、dict文件、json文件。
3. 支持寄存器配置的去重、分组、导出等。

依赖：openpyxl（如需Excel导出）

用法示例：
    from xml_parser import XMLParser
    parser = XMLParser('your.xml')
    parser.parse_to_dict()
    parser.get_regdefing_py()
    parser.xml_to_json()
    parser.write_json_file()

作者：yfzhao

--------------------------------------
"""

import xml.etree.ElementTree as ET
import json
import os
import re
from collections import defaultdict

#build EXCEL file need 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

class XMLParser:
    """
    多功能XML解析器，包含：
    1. 提取设备地址映射功能（原XMLDeviceAddressParser）
    2. XML转Excel功能（原xml_to_excel.py）
    """
    
    def __init__(self, xml_file_path: str):
        """
        初始化解析器
        :param xml_file_path: XML配置文件路径
        """
        self.xml_file = xml_file_path
        self.dev_addr_dict = {}
        self.tree = ET.parse(xml_file_path)
        self.root = self.tree.getroot()
        
        # 自动生成输出文件名
        base_name = os.path.splitext(os.path.basename(xml_file_path))[0]
        self.dict_output_file = f"{base_name}_dev_addr.dict"
        self.pydef_output_file = f"{base_name}_reg_def.py"
        self.excel_output_file = f"{base_name}_Register_Config.xlsx"
        self.json_output_file = f"{base_name}_Register_Config.json"

        #json data init
        self.json_data = {}
    
    # ========== 原XMLDeviceAddressParser功能 ==========
    def parse_to_dict(self) -> dict:
        """
        解析XML文件并返回地址字典
        :return: 包含页面名称到地址映射的字典
        """
        # 按照正确结构解析：file -> device -> interface
        file_node = self.root
        if file_node.tag != 'file':
            raise ValueError("根节点不是file，请检查XML结构")
        
        device = file_node.find('device')
        if device is None:
            raise ValueError("XML文件中未找到device节点")
        
        interfaces = device.findall('interface')
        if not interfaces:
            raise ValueError("XML文件中未找到interface节点")
        
        for interface in interfaces:
            # 获取interface下的name子节点的文本内容
            name_node = interface.find('name')
            if name_node is None:
                continue
                
            page_name = name_node.text
            if not page_name:
                continue
                
            # 获取第一个field节点的address属性
            field = interface.find('field')
            if field is not None:
                address_node = field.find('address')
                if address_node is not None:
                    full_address = address_node.text
                    if full_address:
                        page_address = full_address[:4]  # 取前4位作为页地址
                        self.dev_addr_dict[page_name] = page_address
                        #print(f"{page_name}=={page_address}")
        
        # 生成反向字典，便于地址查找
        self.addr_to_key = {int(v, 16): k for k, v in self.dev_addr_dict.items()}

        return self.dev_addr_dict
    
    def get_page_name_addr(self, output_file: str = None) -> str:
        """
        将解析结果保存为JSON文件
        :param output_file: (可选)自定义输出文件路径
        :return: 实际使用的输出文件路径
        """
        if not self.dev_addr_dict:
            self.parse_to_dict()
            
        output_path = output_file if output_file else self.dict_output_file
            
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.dev_addr_dict, f, ensure_ascii=False, indent=2)
        
        print(f"Write to: {output_path}")
        return output_path

    # ========== copy from pc_conv_aves class ==========
    #output regdefine.py for hier py import
    def get_regdefing_py(self, output_file: str = None) -> str:
        if not self.dev_addr_dict:
            self.parse_to_dict()

        reg_define_lines = []
        for key, value in self.dev_addr_dict.items():
            reg_define_line = f"{key} = {value}"
            reg_define_lines.append(reg_define_line)

        output_path = output_file if output_file else self.pydef_output_file

        with open(output_path, 'w', encoding='utf-8') as file:
            file.write("\n".join(reg_define_lines))
        
        print(f"成功生成reg_define.py文件: {output_path}")
        return output_path

    # ========== 原xml_to_excel.py功能 ==========
    def _parse_mask_shift(self, mask_shift_str):
        """解析mask和shift字段的跨字节配置"""
        if not mask_shift_str or mask_shift_str == "{}":
            return []
        
        entries = []
        # 去除大括号并按逗号分割
        items = mask_shift_str.strip("{}").split(",")
        for item in items:
            addr_part, value_part = item.split(":")
            addr = addr_part.strip()
            value = value_part.strip()
            entries.append((addr, value))
        return entries

    def xml_to_excel(self, excel_file: str = None) -> str:
        """
        将XML转换为Excel文件
        :param excel_file: (可选)自定义输出文件路径
        :return: 实际使用的输出文件路径
        """
        output_path = excel_file if excel_file else self.excel_output_file
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Registers"

        # 增强的标题行
        headers = [
            "Register Name", "Base Address", "Field Name", "Bit Range", 
            "Total Bits", "Default Value", "Data Type", "Description",
            "Byte Address", "Byte Mask", "Byte Shift", "Effective Bits",
            "Configuration Notes"
        ]
        
        # 设置标题样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))

        # 写入标题
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        row_num = 2  # 数据从第2行开始

        for field in self.root.findall(".//field[@class='Field']"):
            name = field.findtext('name')
            caption = field.findtext('caption') or name
            address = field.findtext('address')
            default_value = field.findtext('defaultvalue')
            size = field.findtext('size')
            description = field.findtext('description') or ""
            datatype = field.findtext('datatype') or ""
            mask = field.findtext('mask') or "{}"
            shift = field.findtext('shift') or "{}"
            byteorder = field.findtext('byteorder') or ""

            # 解析基地址
            base_addr = "0x0000"
            if address:
                match = re.match(r"(0x[0-9A-Fa-f]+)", address)
                base_addr = match.group(1) if match else address.split(".")[0]

            # 解析位域范围
            bit_range = None
            range_match = re.search(r"\[(\d+):(\d+)\]", name)
            if range_match:
                msb = int(range_match.group(1))
                lsb = int(range_match.group(2))
                bit_range = f"[{msb}:{lsb}]"
                total_bits = msb - lsb + 1
            else:
                total_bits = int(size) if size else 1

            # 解析mask和shift配置
            mask_entries = self._parse_mask_shift(mask)
            shift_entries = self._parse_mask_shift(shift)
            
            # 合并配置信息
            configs = defaultdict(dict)
            for addr, mask_val in mask_entries:
                configs[addr]["mask"] = mask_val
            for addr, shift_val in shift_entries:
                configs[addr]["shift"] = shift_val

            # 为每个字节配置创建一行
            for byte_addr, byte_config in configs.items():
                # 计算有效位数
                mask_val = byte_config.get("mask", "0x00")
                effective_bits = bin(int(mask_val, 16)).count("1")
                
                # 配置说明
                notes = []
                if byte_config.get("shift"):
                    shift_val = int(byte_config["shift"])
                    if shift_val > 0:
                        notes.append(f"左移{shift_val}位")
                    elif shift_val < 0:
                        notes.append(f"右移{abs(shift_val)}位")
                
                if byteorder == "littleendian":
                    notes.append("小端字节序")
                elif byteorder == "bigendian":
                    notes.append("大端字节序")
                    
                # 写入Excel行数据
                ws.cell(row=row_num, column=1, value=caption)  # Register Name
                ws.cell(row=row_num, column=2, value=base_addr)  # Base Address
                ws.cell(row=row_num, column=3, value=name)  # Field Name
                ws.cell(row=row_num, column=4, value=bit_range)  # Bit Range
                ws.cell(row=row_num, column=5, value=total_bits)  # Total Bits
                ws.cell(row=row_num, column=6, value=default_value)  # Default Value
                ws.cell(row=row_num, column=7, value=datatype)  # Data Type
                ws.cell(row=row_num, column=8, value=description)  # Description
                ws.cell(row=row_num, column=9, value=byte_addr)  # Byte Address
                ws.cell(row=row_num, column=10, value=byte_config.get("mask", ""))  # Byte Mask
                ws.cell(row=row_num, column=11, value=byte_config.get("shift", ""))  # Byte Shift
                ws.cell(row=row_num, column=12, value=effective_bits)  # Effective Bits
                ws.cell(row=row_num, column=13, value="; ".join(notes))  # Configuration Notes

                # 设置数据行样式
                for col_num in range(1, 14):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = thin_border
                    if col_num in [2,4,5,9,10,11,12]:  # 数值列居中
                        cell.alignment = Alignment(horizontal="center")

                row_num += 1

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # 冻结标题行并添加筛选
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

        # 保存Excel文件
        wb.save(output_path)
        print(f"成功生成增强版Excel文件: {output_path}，共 {row_num-2} 行配置数据")
        return output_path

    def _get_key_by_addr(self, addr: int) -> str:
        # 确保dev_addr_dict已初始化
        if not self.dev_addr_dict:
            self.parse_to_dict()
        
        page_name = self.addr_to_key.get(addr)

        if page_name is None:
            page_name = f"0x{addr:02X}" # 如果地址未找到，返回str

        
        return page_name

        

    def _get_base_key(self, byte_address) -> str:
        addr_int = int(byte_address, 16)
        base_addr = addr_int >> 8
        base_name = self._get_key_by_addr(base_addr)
        return base_name
    
    def _organize_registers(self, register_list) -> dict:
        organized = {}
        for reg in register_list:
            byte_addr = reg.get("byte_address")
            if not byte_addr:
                continue
            base_key = self._get_base_key(byte_addr)
            if base_key not in organized:
                organized[base_key] = []
            organized[base_key].append(reg)
        return organized

    def __clear_reg_str(self, str):
        #ADD RULE HERE
        #for strange i2c name
        outstr = str.replace(' ', '_')
        outstr = outstr.replace('/', '_')
        outstr = outstr.replace('-', '_')
        outstr = outstr.replace('[', '_')
        outstr = outstr.replace(']', '_')
        outstr = outstr.replace('(', '_')
        outstr = outstr.replace(')', '_')
        return outstr

    def _clean_reg_name(self):
        """
        在self.json里面清理register_name中的空格
        """
        for key, registers in self.json_data.items():
            for reg in registers:
                reg_name = reg.get("register_name")
                if reg_name:
                    reg["register_name"] = self.__clear_reg_str(reg_name)

    def write_json_file(self, json_file: str = None) -> None:
        """
        将寄存器配置写入JSON文件
        :param json_file: (可选)自定义输出文件路径
        """
        # 如果self.json_data为空，先调用xml_to_json生成数据
        if not self.json_data:
            self.xml_to_json()
        output_path = json_file if json_file else self.json_output_file
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.json_data, f, ensure_ascii=False, indent=4)
        print(f"成功生成JSON文件: {output_path}")

    def xml_to_json(self) -> None:
        """
        将XML转换为JSON文件，包含与to_excel相同的寄存器配置信息
        :param json_file: (可选)自定义输出文件路径
        :return: 实际使用的输出文件路径
        """
        #output_path = json_file if json_file else self.json_output_file
        
        registers = []
        
        for field in self.root.findall(".//field[@class='Field']"):
            name = field.findtext('name')
            caption = field.findtext('caption') or name
            address = field.findtext('address')
            default_value = field.findtext('defaultvalue')
            size = field.findtext('size')
            description = field.findtext('description') or ""
            datatype = field.findtext('datatype') or ""
            mask = field.findtext('mask') or "{}"
            shift = field.findtext('shift') or "{}"
            byteorder = field.findtext('byteorder') or ""

            # 解析基地址
            base_addr = "0x0000"
            if address:
                match = re.match(r"(0x[0-9A-Fa-f]+)", address)
                base_addr = match.group(1) if match else address.split(".")[0]

            # 解析位域范围
            bit_range = None
            range_match = re.search(r"\[(\d+):(\d+)\]", name)
            if range_match:
                msb = int(range_match.group(1))
                lsb = int(range_match.group(2))
                bit_range = f"[{msb}:{lsb}]"
                total_bits = msb - lsb + 1
            else:
                total_bits = int(size) if size else 1

            # 解析mask和shift配置
            mask_entries = self._parse_mask_shift(mask)
            shift_entries = self._parse_mask_shift(shift)
            
            # 合并配置信息
            configs = defaultdict(dict)
            for addr, mask_val in mask_entries:
                configs[addr]["mask"] = mask_val
            for addr, shift_val in shift_entries:
                configs[addr]["shift"] = shift_val

            # 为每个字节配置创建条目
            for byte_addr, byte_config in configs.items():
                # 计算有效位数
                mask_val = byte_config.get("mask", "0x00")
                effective_bits = bin(int(mask_val, 16)).count("1")
                
                # 配置说明
                notes = []
                if byte_config.get("shift"):
                    shift_val = int(byte_config["shift"])
                    if shift_val > 0:
                        notes.append(f"左移{shift_val}位")
                    elif shift_val < 0:
                        notes.append(f"右移{abs(shift_val)}位")
                
                if byteorder == "littleendian":
                    notes.append("小端字节序")
                elif byteorder == "bigendian":
                    notes.append("大端字节序")
                    
                register_data = {
                    "register_name": caption,
                    "base_address": base_addr,
                    "field_name": name,
                    "bit_range": bit_range,
                    "total_bits": total_bits,
                    "default_value": default_value,
                    "data_type": datatype,
                    "description": description,
                    "byte_address": byte_addr,
                    "byte_mask": byte_config.get("mask", ""),
                    "byte_shift": byte_config.get("shift", ""),
                    "effective_bits": effective_bits,
                    "configuration_notes": "; ".join(notes)
                }
                
                registers.append(register_data)

        # 组织寄存器数据, 可选是否write to json
        self.json_data = self._organize_registers(registers)
        # 清理寄存器名称中的空格
        self._clean_reg_name()

        return



# 使用示例
if __name__ == "__main__":
    xml_file = "d:/GS_Projects/gs_svn/gsu1001/eval/aves/gsu1001/GSU1K1_R3.xml"
    #xml_file = "h:/gs_svn/gsu1001/eval/aves/gsu1001/GSU1K1_R3.xml"
    
    # 创建解析器实例
    parser = XMLParser(xml_file)
    
    #parser.get_page_name_addr()  # 自动保存为"GSU1K1_R3_dev_addr.dict"
    
    # STEP 1: 提取设备地址并保存为PY文件
    #parser.get_regdefing_py()  # 自动保存为"GSU1K1_R3_reg_def.py"
    
    # STEP 2, OPTION: 转换为Excel文件
    #parser.xml_to_excel()  # 自动保存为"GSU1K1_R3_Register_Config.xlsx"

    # STEP 3: 转换为JSON文件
    parser.write_json_file()  # 自动保存为"GSU1K1_R3_Register_Config.json"